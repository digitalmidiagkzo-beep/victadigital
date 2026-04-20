"""
VictaDigital — Gerador de páginas Hugo
Lê a planilha VictaDigital_Planilha.xlsx e gera os arquivos .md automaticamente.

Como usar:
1. Coloque este script na pasta do projeto: C:/Users/gleic/Documents/victadigital/
2. Coloque a planilha na mesma pasta
3. No terminal do VS Code, rode: python gerar_paginas.py
"""

import os
import re

# ── Tenta importar openpyxl, instala se não tiver ──────────────────────────
try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

# ── Configurações ───────────────────────────────────────────────────────────
PLANILHA = "VictaDigital_Planilha.xlsx"
ABA      = "Negócios"
PASTA_CONTENT = "content"

# ── Funções utilitárias ─────────────────────────────────────────────────────
def slugify(texto):
    """Converte texto para slug sem acentos."""
    texto = str(texto).lower().strip()
    subs = {
        'á':'a','à':'a','ã':'a','â':'a','ä':'a',
        'é':'e','è':'e','ê':'e','ë':'e',
        'í':'i','ì':'i','î':'i','ï':'i',
        'ó':'o','ò':'o','õ':'o','ô':'o','ö':'o',
        'ú':'u','ù':'u','û':'u','ü':'u',
        'ç':'c','ñ':'n','ý':'y',
    }
    for k, v in subs.items():
        texto = texto.replace(k, v)
    texto = re.sub(r'[^a-z0-9\s-]', '', texto)
    texto = re.sub(r'[\s]+', '-', texto)
    texto = re.sub(r'-+', '-', texto)
    return texto.strip('-')

def val(cell):
    """Retorna valor da célula como string limpa."""
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()

def gerar_slug_negocio(categoria, cidade, bairro, nome):
    """Gera slug no padrão VictaDigital."""
    cat_slug  = slugify(categoria)
    cid_slug  = slugify(cidade)
    bai_slug  = slugify(bairro)
    nome_slug = slugify(nome)
    # Ex: dentistas/dentista-contagem-eldorado/clinica-sorriso
    pasta    = f"{cat_slug}/{cat_slug[:-1] if cat_slug.endswith('s') else cat_slug}-{cid_slug}-{bai_slug}"
    arquivo  = nome_slug
    return pasta, arquivo

def gerar_md_negocio(dados):
    """Gera o conteúdo do arquivo .md para página individual."""
    whatsapp    = dados.get('whatsapp', '')
    patrocinado = dados.get('patrocinado', 'false').lower()
    cta = dados.get('cta_text') or 'Quero conhecer mais sobre os serviços'

    # Gera title e meta automaticamente se não preenchido
    nome    = dados.get('nome_negocio', '')
    bairro  = dados.get('bairro', '')
    cidade  = dados.get('cidade', '')
    cat     = dados.get('categoria', '')
    title   = f"{nome} | {cat} no {bairro}, {cidade}"
    meta    = f"Conheça {nome}, especialista em {cat.lower()} no {bairro}, {cidade}. Agendamento direto pelo WhatsApp."

    servicos = dados.get('servicos_lista', '')
    bv       = dados.get('bairros_vizinhos', '')
    endereco = dados.get('endereco', '')
    avaliacao = dados.get('avaliacao', '4.9')
    num_av    = dados.get('num_avaliacoes', '47')
    slug_completo = dados.get('slug', '')

    # Slug = apenas o último segmento do caminho
    slug_final = slug_completo.split('/')[-1] if slug_completo else slugify(nome)

    linhas = [
        "---",
        f'title: "{title}"',
        f'meta_description: "{meta}"',
        f'slug: {slug_final}',
        f'servico: {cat}',
        f'nome_cliente: {nome}',
        f'cidade: {cidade}',
        f'bairro: {bairro}',
        f'avaliacao: "{avaliacao}"',
        f'num_avaliacoes: "{num_av}"',
    ]

    if endereco:
        linhas.append(f'endereco: "{endereco}"')

    if whatsapp:
        linhas.append(f'whatsapp: "{whatsapp}"')
        linhas.append(f'cta_text: "{cta}"')
    
    if dados.get('foto_cliente'):
        linhas.append(f'foto_cliente: "{dados["foto_cliente"]}"')

    if servicos:
        linhas.append(f'servicos_lista: "{servicos}"')

    if bv:
        linhas.append(f'bairros_vizinhos: "{bv}"')

    linhas.append(f'patrocinado: {patrocinado}')
    linhas.append('layout: pseo')
    linhas.append("---")

    return "\n".join(linhas)

# ── Leitura da planilha ─────────────────────────────────────────────────────
print(f"\n{'='*50}")
print("VictaDigital — Gerador de Páginas Hugo")
print(f"{'='*50}\n")

if not os.path.exists(PLANILHA):
    print(f"ERRO: Planilha '{PLANILHA}' não encontrada.")
    print("Certifique-se que a planilha está na mesma pasta que este script.")
    input("\nPressione Enter para sair...")
    exit()

wb = openpyxl.load_workbook(PLANILHA)

if ABA not in wb.sheetnames:
    print(f"ERRO: Aba '{ABA}' não encontrada na planilha.")
    input("\nPressione Enter para sair...")
    exit()

ws = wb[ABA]

# Mapeamento das colunas (linha 4 é o cabeçalho)
colunas = {}
for col in ws.iter_cols(min_row=4, max_row=4):
    header = val(col[0]).replace("\n", " ").strip()
    colunas[header] = col[0].column

print(f"Colunas encontradas: {list(colunas.keys())}\n")

# ── Mapeamento simplificado por posição ────────────────────────────────────
# A=1 cidade, B=2 categoria, C=3 bairro, D=4 nome, E=5 slug
# F=6 endereço, G=7 bairros_viz, H=8 lat, I=9 lng
# J=10 avaliacao, K=11 num_av, L=12 servicos, M=13 horario
# N=14 whatsapp, O=15 foto, P=16 cta, Q=17 patrocinado
# R=18 status, S=19 obs

COL = {
    'cidade': 1, 'categoria': 2, 'bairro': 3,
    'nome_negocio': 4, 'slug': 5,
    'endereco': 6, 'bairros_vizinhos': 7,
    'avaliacao': 10, 'num_avaliacoes': 11,
    'servicos_lista': 12,
    'whatsapp': 14, 'foto_cliente': 15,
    'cta_text': 16, 'patrocinado': 17,
    'status': 18,
}

criados   = 0
pulados   = 0
erros     = 0

# Processa a partir da linha 5 (dados começam na linha 5)
for row_num in range(5, ws.max_row + 1):
    cidade   = val(ws.cell(row=row_num, column=COL['cidade']))
    nome     = val(ws.cell(row=row_num, column=COL['nome_negocio']))
    categoria = val(ws.cell(row=row_num, column=COL['categoria']))
    bairro   = val(ws.cell(row=row_num, column=COL['bairro']))
    status   = val(ws.cell(row=row_num, column=COL['status'])).lower()

    # Pula linhas vazias ou com status "sem interesse"
    if not nome or not cidade:
        continue
    if status == "sem interesse":
        print(f"  ⏭  Pulando '{nome}' (sem interesse)")
        pulados += 1
        continue

    # Lê todos os dados
    dados = {}
    for campo, col_idx in COL.items():
        dados[campo] = val(ws.cell(row=row_num, column=col_idx))

    # Gera slug se não preenchido
    slug = dados.get('slug', '').strip()
    if not slug:
        pasta_slug, nome_slug = gerar_slug_negocio(categoria, cidade, bairro, nome)
        slug = f"{pasta_slug}/{nome_slug}"
        dados['slug'] = slug

    # Define caminho do arquivo
    partes = slug.split('/')
    if len(partes) < 2:
        print(f"  ❌ Slug inválido para '{nome}': {slug}")
        erros += 1
        continue

    # Cada negócio vira uma pasta com index.md dentro
    caminho_pasta = os.path.join(PASTA_CONTENT, *partes)
    caminho_md = os.path.join(caminho_pasta, "index.md")

    # Cria pasta se não existir
    os.makedirs(caminho_pasta, exist_ok=True)

    # Verifica se já existe
    if os.path.exists(caminho_md):
        print(f"  ⚠️  Já existe: {caminho_md} (pulando)")
        pulados += 1
        continue

    # Gera conteúdo e salva
    conteudo = gerar_md_negocio(dados)
    with open(caminho_md, 'w', encoding='utf-8') as f:
        f.write(conteudo)

    print(f"  ✅ Criado: {caminho_md}")
    criados += 1

# ── Resumo ──────────────────────────────────────────────────────────────────
print(f"\n{'='*50}")
print(f"✅ Criados:  {criados} arquivos")
print(f"⏭  Pulados:  {pulados} linhas")
print(f"❌ Erros:    {erros}")
print(f"{'='*50}")
print("\nAgora rode no terminal: hugo server")
print("E verifique as páginas no browser!\n")

input("Pressione Enter para fechar...")
